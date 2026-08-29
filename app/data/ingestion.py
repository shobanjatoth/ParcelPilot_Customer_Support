# import os
# import json
# import uuid
# import pymupdf
# import openpyxl
# from sqlalchemy.orm import Session
# from app.data.models import Account, Order, Ticket
# from app.vector.store import VectorStore
# from app.services.reliability import DOCUMENT_METADATA

# RAW_DIR = os.path.join(os.path.dirname(__file__), "..", "..", "data", "raw")


# def _read_readme_timestamp(wb) -> str | None:
#     for name in wb.sheetnames:
#         if name.lower() in ("readme", "read me", "info", "metadata"):
#             sheet = wb[name]
#             for row in sheet.iter_rows(values_only=True):
#                 for cell in row:
#                     if cell and isinstance(cell, str) and "snapshot" in cell.lower():
#                         idx = row.index(cell)
#                         if idx + 1 < len(row) and row[idx + 1]:
#                             return str(row[idx + 1])
#                     if cell and isinstance(cell, str) and "timestamp" in cell.lower():
#                         idx = row.index(cell)
#                         if idx + 1 < len(row) and row[idx + 1]:
#                             return str(row[idx + 1])
#     return None


# def _find_sheet(wb, preferred_names: list[str]) -> str | None:
#     for name in preferred_names:
#         if name in wb.sheetnames:
#             return name
#     for sheet_name in wb.sheetnames:
#         lower = sheet_name.lower()
#         for pref in preferred_names:
#             if pref.lower() in lower:
#                 return sheet_name
#     return None


# def ingest_excel(db: Session):
#     path = os.path.join(RAW_DIR, "ParcelPilot_Assessment_Data.xlsx")
#     wb = openpyxl.load_workbook(path, data_only=True)

#     snapshot_time = _read_readme_timestamp(wb)
#     if snapshot_time:
#         print(f"  README snapshot timestamp: {snapshot_time}")

#     accounts_sheet_name = _find_sheet(wb, ["accounts", "account"])
#     if accounts_sheet_name:
#         accounts_sheet = wb[accounts_sheet_name]
#         headers = [cell.value for cell in accounts_sheet[1]]
#         for row in accounts_sheet.iter_rows(min_row=2, values_only=True):
#             data = dict(zip(headers, row))
#             account = Account(
#                 account_id=str(data["account_id"]),
#                 account_name=str(data["account_name"]),
#                 plan=str(data["plan"]),
#                 status=str(data["status"]),
#                 csm=str(data["csm"]) if data.get("csm") else None,
#                 contract_file=str(data["contract_file"]) if data.get("contract_file") else None,
#                 premium_support=bool(data.get("premium_support", False)),
#                 notes=str(data["notes"]) if data.get("notes") else None,
#             )
#             db.merge(account)
#         print(f"  Ingested accounts from sheet: {accounts_sheet_name}")

#     orders_sheet_name = _find_sheet(wb, ["orders", "order"])
#     if orders_sheet_name:
#         orders_sheet = wb[orders_sheet_name]
#         headers = [cell.value for cell in orders_sheet[1]]
#         for row in orders_sheet.iter_rows(min_row=2, values_only=True):
#             data = dict(zip(headers, row))
#             order = Order(
#                 order_id=str(data["order_id"]),
#                 account_id=str(data["account_id"]),
#                 carrier=str(data["carrier"]),
#                 status=str(data["status"]),
#                 booked_at=str(data["booked_at"]) if data.get("booked_at") else None,
#                 pickup_window_start=str(data["pickup_window_start"]) if data.get("pickup_window_start") else None,
#                 pickup_window_end=str(data["pickup_window_end"]) if data.get("pickup_window_end") else None,
#                 pickup_actual_at=str(data["pickup_actual_at"]) if data.get("pickup_actual_at") else None,
#                 shipment_fee_inr=float(data["shipment_fee_inr"]) if data.get("shipment_fee_inr") else 0.0,
#                 carrier_fault=bool(data.get("carrier_fault", False)),
#                 customer_fault=bool(data.get("customer_fault", False)),
#                 cancellation_requested_at=str(data["cancellation_requested_at"]) if data.get("cancellation_requested_at") else None,
#                 notes=str(data["notes"]) if data.get("notes") else None,
#             )
#             db.merge(order)
#         print(f"  Ingested orders from sheet: {orders_sheet_name}")

#     tickets_sheet_name = _find_sheet(wb, ["tickets", "ticket"])
#     if tickets_sheet_name:
#         tickets_sheet = wb[tickets_sheet_name]
#         headers = [cell.value for cell in tickets_sheet[1]]
#         for row in tickets_sheet.iter_rows(min_row=2, values_only=True):
#             data = dict(zip(headers, row))
#             ticket = Ticket(
#                 ticket_id=str(data["ticket_id"]),
#                 account_id=str(data["account_id"]),
#                 created_at=str(data["created_at"]) if data.get("created_at") else None,
#                 status=str(data["status"]),
#                 subject=str(data["subject"]) if data.get("subject") else None,
#                 description=str(data["description"]) if data.get("description") else None,
#                 channel=str(data["channel"]) if data.get("channel") else None,
#                 assigned_to=str(data["assigned_to"]) if data.get("assigned_to") else None,
#                 last_customer_message_at=str(data["last_customer_message_at"]) if data.get("last_customer_message_at") else None,
#                 historical_resolution=str(data["historical_resolution"]) if data.get("historical_resolution") else None,
#             )
#             db.merge(ticket)
#         print(f"  Ingested tickets from sheet: {tickets_sheet_name}")

#     db.commit()
#     wb.close()


# def ingest_pdfs(vector_store: VectorStore):
#     pdf_files = [
#         "01_Support_Policy_v3_CURRENT.pdf",
#         "02_Support_Policy_v2_DEPRECATED.pdf",
#         "03_Cancellation_and_Service_Credit_SOP_v4.pdf",
#         "04_Product_Operations_Guide_and_Known_Issues.pdf",
#         "05_Northstar_Logistics_Enterprise_Agreement.pdf",
#         "06_LumenWorks_Service_Agreement.pdf",
#     ]

#     for pdf_file in pdf_files:
#         path = os.path.join(RAW_DIR, pdf_file)
#         if not os.path.exists(path):
#             print(f"  Skipping {pdf_file} - not found")
#             continue

#         meta = DOCUMENT_METADATA.get(pdf_file, {})
#         doc = pymupdf.open(path)
#         chunks = []

#         for page_num in range(len(doc)):
#             page = doc[page_num]
#             text = page.get_text()
#             if not text.strip():
#                 continue

#             paragraphs = [p.strip() for p in text.split("\n\n") if p.strip()]
#             if not paragraphs:
#                 paragraphs = [text.strip()]

#             for i, paragraph in enumerate(paragraphs):
#                 if len(paragraph) < 10:
#                     continue
#                 chunk_id = f"{pdf_file}:p{page_num + 1}:c{i}"
#                 chunks.append({
#                     "id": chunk_id,
#                     "text": paragraph,
#                     "metadata": {
#                         "document_name": meta.get("document_name", pdf_file),
#                         "document_type": meta.get("document_type", "unknown"),
#                         "version": meta.get("version", "unknown"),
#                         "status": meta.get("status", "unknown"),
#                         "effective_date": meta.get("effective_date", "unknown"),
#                         "customer_account_id": meta.get("customer_account_id"),
#                         "source_priority": meta.get("source_priority", 50),
#                         "page_number": page_num + 1,
#                         "section": meta.get("section", "general"),
#                         "source_file": pdf_file,
#                     },
#                 })

#         doc.close()

#         if chunks:
#             vector_store.add_documents(chunks)
#             print(f"  Ingested {pdf_file}: {len(chunks)} chunks")


# def run_ingestion(db: Session, vector_store: VectorStore):
#     print("Ingesting Excel data...")
#     ingest_excel(db)
#     print("Excel ingestion complete.")

#     print("Ingesting PDF documents...")
#     ingest_pdfs(vector_store)
#     print("PDF ingestion complete.")


from __future__ import annotations

import os
import uuid
from typing import Any

import openpyxl
import pymupdf
from sqlalchemy.orm import Session

from app.data.models import Account, Order, Ticket
from app.services.reliability import DOCUMENT_METADATA
from app.vector.store import VectorStore


# ============================================================
# Paths
# ============================================================

BASE_DIR = os.path.abspath(
    os.path.join(
        os.path.dirname(__file__),
        "..",
        "..",
    )
)

RAW_DIR = os.path.join(
    BASE_DIR,
    "data",
    "raw",
)


EXCEL_FILE = "ParcelPilot_Assessment_Data.xlsx"


PDF_FILES = [
    "01_Support_Policy_v3_CURRENT.pdf",
    "02_Support_Policy_v2_DEPRECATED.pdf",
    "03_Cancellation_and_Service_Credit_SOP_v4.pdf",
    "04_Product_Operations_Guide_and_Known_Issues.pdf",
    "05_Northstar_Logistics_Enterprise_Agreement.pdf",
    "06_LumenWorks_Service_Agreement.pdf",
]


# ============================================================
# Excel helpers
# ============================================================

def _find_sheet(
    workbook,
    preferred_names: list[str],
) -> str | None:

    # Exact match first
    for name in preferred_names:
        if name in workbook.sheetnames:
            return name

    # Partial match
    for sheet_name in workbook.sheetnames:
        lower_name = sheet_name.lower()

        for preferred in preferred_names:
            if preferred.lower() in lower_name:
                return sheet_name

    return None


def _read_sheet_rows(sheet) -> list[dict[str, Any]]:

    rows = list(
        sheet.iter_rows(
            values_only=True
        )
    )

    if not rows:
        return []

    headers = [
        str(header).strip()
        if header is not None
        else ""
        for header in rows[0]
    ]

    result = []

    for row in rows[1:]:

        if not any(value is not None for value in row):
            continue

        data = dict(
            zip(
                headers,
                row,
            )
        )

        result.append(data)

    return result


# ============================================================
# Excel → PostgreSQL
# ============================================================

def ingest_excel(db: Session) -> None:

    path = os.path.join(
        RAW_DIR,
        EXCEL_FILE,
    )

    if not os.path.exists(path):
        raise FileNotFoundError(
            f"Excel file not found: {path}"
        )

    print(f"Reading Excel: {path}")

    workbook = openpyxl.load_workbook(
        path,
        data_only=True,
    )

    # --------------------------------------------------------
    # Accounts
    # --------------------------------------------------------

    accounts_sheet = _find_sheet(
        workbook,
        ["accounts", "account"],
    )

    if accounts_sheet:

        rows = _read_sheet_rows(
            workbook[accounts_sheet]
        )

        for data in rows:

            account_id = data.get("account_id")

            if not account_id:
                continue

            account = Account(
                account_id=str(account_id),
                account_name=str(
                    data.get("account_name", "")
                ),
                plan=str(
                    data.get("plan", "")
                ),
                status=str(
                    data.get("status", "active")
                ),
                csm=(
                    str(data["csm"])
                    if data.get("csm")
                    else None
                ),
                contract_file=(
                    str(data["contract_file"])
                    if data.get("contract_file")
                    else None
                ),
                premium_support=bool(
                    data.get(
                        "premium_support",
                        False,
                    )
                ),
                notes=(
                    str(data["notes"])
                    if data.get("notes")
                    else None
                ),
            )

            db.merge(account)

        print(
            f"Accounts ingested: {len(rows)}"
        )

    # --------------------------------------------------------
    # Orders
    # --------------------------------------------------------

    orders_sheet = _find_sheet(
        workbook,
        ["orders", "order"],
    )

    if orders_sheet:

        rows = _read_sheet_rows(
            workbook[orders_sheet]
        )

        for data in rows:

            order_id = data.get("order_id")

            if not order_id:
                continue

            order = Order(
                order_id=str(order_id),
                account_id=str(
                    data.get("account_id", "")
                ),
                carrier=str(
                    data.get("carrier", "")
                ),
                status=str(
                    data.get("status", "")
                ),
                booked_at=_string_or_none(
                    data.get("booked_at")
                ),
                pickup_window_start=_string_or_none(
                    data.get("pickup_window_start")
                ),
                pickup_window_end=_string_or_none(
                    data.get("pickup_window_end")
                ),
                pickup_actual_at=_string_or_none(
                    data.get("pickup_actual_at")
                ),
                shipment_fee_inr=float(
                    data.get(
                        "shipment_fee_inr",
                        0,
                    ) or 0
                ),
                carrier_fault=bool(
                    data.get(
                        "carrier_fault",
                        False,
                    )
                ),
                customer_fault=bool(
                    data.get(
                        "customer_fault",
                        False,
                    )
                ),
                cancellation_requested_at=_string_or_none(
                    data.get(
                        "cancellation_requested_at"
                    )
                ),
                notes=_string_or_none(
                    data.get("notes")
                ),
            )

            db.merge(order)

        print(
            f"Orders ingested: {len(rows)}"
        )

    # --------------------------------------------------------
    # Tickets
    # --------------------------------------------------------

    tickets_sheet = _find_sheet(
        workbook,
        ["tickets", "ticket"],
    )

    if tickets_sheet:

        rows = _read_sheet_rows(
            workbook[tickets_sheet]
        )

        for data in rows:

            ticket_id = data.get("ticket_id")

            if not ticket_id:
                continue

            ticket = Ticket(
                ticket_id=str(ticket_id),
                account_id=str(
                    data.get("account_id", "")
                ),
                created_at=_string_or_none(
                    data.get("created_at")
                ),
                status=str(
                    data.get("status", "open")
                ),
                subject=_string_or_none(
                    data.get("subject")
                ),
                description=_string_or_none(
                    data.get("description")
                ),
                channel=_string_or_none(
                    data.get("channel")
                ),
                assigned_to=_string_or_none(
                    data.get("assigned_to")
                ),
                last_customer_message_at=_string_or_none(
                    data.get(
                        "last_customer_message_at"
                    )
                ),
                historical_resolution=_string_or_none(
                    data.get(
                        "historical_resolution"
                    )
                ),
            )

            db.merge(ticket)

        print(
            f"Tickets ingested: {len(rows)}"
        )

    db.commit()

    workbook.close()

    print("Excel → PostgreSQL complete.")


# ============================================================
# Utility
# ============================================================

def _string_or_none(
    value: Any,
) -> str | None:

    if value is None:
        return None

    return str(value)


# ============================================================
# PDF chunking
# ============================================================

def _chunk_text(
    text: str,
    max_chars: int = 1500,
    overlap: int = 200,
) -> list[str]:

    text = text.strip()

    if not text:
        return []

    paragraphs = [
        paragraph.strip()
        for paragraph in text.split("\n\n")
        if paragraph.strip()
    ]

    chunks: list[str] = []

    current = ""

    for paragraph in paragraphs:

        if len(current) + len(paragraph) + 2 <= max_chars:

            if current:
                current += "\n\n"

            current += paragraph

        else:

            if current:
                chunks.append(current)

            # Preserve overlap from previous chunk
            if overlap > 0 and current:
                current = current[
                    max(
                        0,
                        len(current) - overlap,
                    ):
                ]
                current += "\n\n" + paragraph
            else:
                current = paragraph

    if current:
        chunks.append(current)

    return chunks


# ============================================================
# PDFs → Qdrant
# ============================================================

def ingest_pdfs(
    vector_store: VectorStore,
) -> None:

    total_chunks = 0

    for pdf_file in PDF_FILES:

        path = os.path.join(
            RAW_DIR,
            pdf_file,
        )

        if not os.path.exists(path):

            print(
                f"Skipping missing PDF: {pdf_file}"
            )

            continue

        metadata = DOCUMENT_METADATA.get(
            pdf_file,
            {},
        )

        print(
            f"Processing PDF: {pdf_file}"
        )

        document = pymupdf.open(path)

        chunks: list[dict[str, Any]] = []

        for page_index in range(
            len(document)
        ):

            page_number = page_index + 1

            page = document[page_index]

            text = page.get_text()

            if not text.strip():
                continue

            page_chunks = _chunk_text(
                text
            )

            for chunk_index, chunk in enumerate(
                page_chunks
            ):

                if len(chunk.strip()) < 20:
                    continue

                source_id = (
                    f"{pdf_file}:"
                    f"p{page_number}:"
                    f"c{chunk_index}"
                )

                # Deterministic ID
                # means repeated ingestion
                # produces the same Qdrant point.
                chunk_uuid = str(
                    uuid.uuid5(
                        uuid.NAMESPACE_URL,
                        source_id,
                    )
                )

                chunk_metadata = {
                    "document_name": metadata.get(
                        "document_name",
                        pdf_file,
                    ),
                    "document_type": metadata.get(
                        "document_type",
                        "unknown",
                    ),
                    "version": metadata.get(
                        "version",
                        "unknown",
                    ),
                    "status": metadata.get(
                        "status",
                        "unknown",
                    ),
                    "effective_date": metadata.get(
                        "effective_date",
                        "unknown",
                    ),
                    "customer_account_id": metadata.get(
                        "customer_account_id"
                    ),
                    "source_priority": metadata.get(
                        "source_priority",
                        50,
                    ),
                    "section": metadata.get(
                        "section",
                        "general",
                    ),
                    "page_number": page_number,
                    "source_file": pdf_file,
                    "chunk_index": chunk_index,
                }

                chunks.append(
                    {
                        "id": chunk_uuid,
                        "text": chunk,
                        "metadata": chunk_metadata,
                    }
                )

        document.close()

        if chunks:

            vector_store.add_documents(
                chunks
            )

            total_chunks += len(chunks)

            print(
                f"  Added {len(chunks)} chunks"
            )

    print(
        f"PDF → Qdrant complete. "
        f"Total chunks: {total_chunks}"
    )


# ============================================================
# Complete ingestion
# ============================================================

def run_ingestion(
    db: Session,
    vector_store: VectorStore,
) -> None:

    print("=" * 60)
    print("PARCELPILOT DATA INGESTION")
    print("=" * 60)

    print("\n[1/2] Excel → PostgreSQL")

    ingest_excel(db)

    print("\n[2/2] PDFs → Qdrant Cloud")

    ingest_pdfs(vector_store)

    print("\n" + "=" * 60)
    print("INGESTION COMPLETE")
    print("=" * 60)
